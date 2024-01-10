# -*- coding: utf-8 -*-
"""
Created on Wed Jan 10 08:24:35 2024

@author: ASUS
"""
import streamlit as st
import os
import pandas as pd
import plotly.express as px
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

def can_open_file(file_path):
    try:
        with open(file_path, 'rb'):
            pass
        return True
    except Exception as e:
        print(f"Error opening file {file_path}: {e}")
        return False

def timestamp_to_datetime(timestamp):
    return datetime.fromtimestamp(timestamp).strftime('%d/%m/%y %I:%M %p')

def create_excel_report_local(folder_path, report_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Local Folder Report"
    ws.append(['Nombre de Archivo', 'Tipo de Archivo', 'Tamaño (bytes)', 'Se puede abrir', 'Ultima modificación', 'Subfolder 1', 'Subfolder 2', 'Subfolder 3', 'Subfolder 4', 'Subfolder 5', 'Subfolder 6', 'Subfolder 7', 'Subfolder 8', 'Subfolder 9', 'Subfolder 10'])
    
    for root, dirs, files in os.walk(folder_path):
        # Exclude hidden folders
        dirs[:] = [d for d in dirs if not d.startswith('.')]

        for file_name in files:
            file_path = os.path.join(root, file_name)
            file_size = 0
            file_type = file_name.split('.')[-1] if '.' in file_name else 'Unknown'
            can_open = can_open_file(file_path)

            try:
                file_size = os.path.getsize(file_path)
                last_modified = os.path.getmtime(file_path)
            except Exception as e:
                print(f"Error getting size or last modified for {file_path}: {e}")

            # Split the folder path into subfolders
            subfolders = os.path.relpath(root, folder_path).split(os.path.sep)[:10]

            # Fill in empty values if there are fewer than 10 subfolders
            subfolders.extend([''] * (10 - len(subfolders)))

            last_modified_formatted = timestamp_to_datetime(last_modified)

            ws.append([file_name, file_type, file_size, can_open, last_modified_formatted] + subfolders)

    # Auto-adjust column widths and apply alignment
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
        ws.column_dimensions[column].alignment = Alignment(horizontal='left', vertical='center')

    wb.save(report_path)



def main():
    st.title("Generador de Reporte")

    st.header("Generador de Reporte de Carpeta Local")
    
    # Use a relative path from the current working directory
    folder_path_local = st.text_input("Ruta de la carpeta (relativa al directorio de trabajo)", value="proyectos/180M- OK Migrada")
    
    report_path_local = st.text_input("Ruta del Reporte (usar .xlsx al final del nombre del reporte)")

    if st.button("Generar Reporte"):
        # Convert the relative path to an absolute path
        folder_path_local_absolute = os.path.abspath(folder_path_local)
        
        create_excel_report_local(folder_path_local_absolute, report_path_local)
        st.success("Reporte Generado exitosamente!")

        st.write(f"Provided folder path: {folder_path_local_absolute}")
        st.write(f"Is folder accessible? {os.path.exists(folder_path_local_absolute)}")
        
        # Bar plot showing the file types for the local folder report
        st.title("Distribución de archivos (Carpeta Local)")
        if os.path.exists(report_path_local):
            df_local = pd.read_excel(report_path_local)
            total_files_local = len(df_local)  # Total number of files
            st.info(f'Total Number of Files: {total_files_local}')

            file_types_local = df_local['Tipo de Archivo'].apply(lambda x: x.lower().strip() if isinstance(x, str) else 'Unknown').value_counts()
            fig_local = px.bar(df_local, x='Tipo de Archivo', labels={'x': 'Tipo de Archivo', 'y': 'Count'})
            st.plotly_chart(fig_local)

        else:
            st.warning("Generar el reporte para ver distribución de archivos")

if __name__ == '__main__':
    main()

